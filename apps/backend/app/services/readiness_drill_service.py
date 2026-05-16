"""Audit readiness KPI drill-down — rich detail for /kpi/drilldown/audit_readiness_pct."""

from __future__ import annotations

import csv
import io
from typing import Any, Dict, List, Optional, Tuple

from datetime import datetime, timezone

import os
from collections import defaultdict

from app.analytics import _scope_exceptions, cfo_cockpit
from app.services import kpi_snapshot_service as kss
from app.services.case_service import merge_cases_master_filters
from app.utils.timeutil import iso_utc


def readiness_drill_v2_enabled() -> bool:
    return os.getenv("ENABLE_READINESS_DRILL_V2", "true").lower() in ("1", "true", "yes", "on")


def _as_of_now() -> str:
    return iso_utc(datetime.now(timezone.utc))


def _severity_for_pct(v: Optional[float], good: float, warn: float) -> Optional[str]:
    if v is None:
        return None
    if v >= good:
        return "success"
    if v >= warn:
        return "warning"
    return "critical"


READINESS_TARGET_PCT = 80.0
READINESS_WARN_PCT = 60.0

CORRELATED_KPI_IDS = (
    "repeat_finding_rate_pct",
    "evidence_completeness_pct",
    "remediation_sla_pct",
    "high_critical_open_cases",
    "unresolved_high_risk_exposure",
)


def _avg_component(heatmap: List[Dict[str, Any]], key: str) -> float:
    if not heatmap:
        return 0.0
    vals = [float(r.get(key) or 0) for r in heatmap]
    return round(sum(vals) / len(vals) * 100, 1)


def _distribution(heatmap: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    buckets = {"0_60": 0, "60_80": 0, "80_100": 0}
    for r in heatmap:
        v = float(r.get("readiness") or 0)
        if v < 60:
            buckets["0_60"] += 1
        elif v < 80:
            buckets["60_80"] += 1
        else:
            buckets["80_100"] += 1
    total = max(1, len(heatmap))
    return [
        {"bucket": "0–60%", "count": buckets["0_60"], "pct": round(100 * buckets["0_60"] / total, 1)},
        {"bucket": "60–80%", "count": buckets["60_80"], "pct": round(100 * buckets["60_80"] / total, 1)},
        {"bucket": "80–100%", "count": buckets["80_100"], "pct": round(100 * buckets["80_100"] / total, 1)},
    ]


def _weakest_cell(heatmap: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not heatmap:
        return None
    row = min(heatmap, key=lambda r: float(r.get("readiness") or 100))
    return {
        "entity": row.get("entity"),
        "process": row.get("process"),
        "readiness": row.get("readiness"),
        "open_high": row.get("open_high"),
        "exposure": row.get("exposure"),
    }


def _weakest_by_key(heatmap: List[Dict[str, Any]], key: str) -> Optional[Dict[str, Any]]:
    if not heatmap:
        return None
    agg: Dict[str, List[float]] = {}
    for r in heatmap:
        k = str(r.get(key) or "unknown")
        agg.setdefault(k, []).append(float(r.get("readiness") or 0))
    best_name = min(agg.keys(), key=lambda n: sum(agg[n]) / len(agg[n]))
    avg = round(sum(agg[best_name]) / len(agg[best_name]), 1)
    return {key: best_name, "readiness": avg, "cell_count": len(agg[best_name])}


def _portfolio_components(heatmap: List[Dict[str, Any]]) -> Dict[str, Any]:
    return {
        "control_pct": _avg_component(heatmap, "control_component"),
        "recon_pct": _avg_component(heatmap, "recon_component"),
        "evidence_pct": _avg_component(heatmap, "evidence_component"),
        "issue_pct": _avg_component(heatmap, "issue_component"),
        "weights": {"control": 40, "recon": 25, "evidence": 20, "issues": 15},
    }


READINESS_ALERT_HREFS = {
    "audit_readiness_pct": "/app/kpi/audit_readiness_pct",
    "repeat_finding_rate_pct": "/app/kpi/repeat_finding_rate_pct",
    "high_critical_open_cases": "/app/cases?status=open",
    "remediation_sla_pct": "/app/cases",
    "evidence_completeness_pct": "/app/evidence",
}


def _movers_from_heatmap(
    current: List[Dict[str, Any]],
    prior_cells: Optional[Dict[str, float]] = None,
) -> Dict[str, Any]:
    """Week-over-week cell movers when prior snapshot stored readiness_cells."""
    rows_with_delta: List[Dict[str, Any]] = []
    for r in current:
        ent = r.get("entity")
        proc = r.get("process")
        if not ent or not proc:
            continue
        key = f"{ent}|{proc}"
        cur = float(r.get("readiness") or 0)
        prior_val = (prior_cells or {}).get(key) if prior_cells else None
        row = {
            "entity": ent,
            "process": proc,
            "readiness": r.get("readiness"),
            "prior_readiness": prior_val,
            "delta_pts": round(cur - float(prior_val), 1) if prior_val is not None else None,
            "open_high": r.get("open_high"),
            "exposure": r.get("exposure"),
        }
        rows_with_delta.append(row)

    if prior_cells and any(r.get("delta_pts") is not None for r in rows_with_delta):
        with_delta = [r for r in rows_with_delta if r.get("delta_pts") is not None]
        deteriorators = sorted(with_delta, key=lambda x: x["delta_pts"])[:5]
        improvers = sorted(with_delta, key=lambda x: -x["delta_pts"])[:5]
        return {
            "mode": "wow",
            "top_deteriorators": deteriorators,
            "top_improvers": improvers,
        }

    deteriorators = sorted(rows_with_delta, key=lambda x: float(x.get("readiness") or 100))[:5]
    improvers = sorted(rows_with_delta, key=lambda x: -float(x.get("readiness") or 0))[:5]
    return {
        "mode": "level",
        "top_deteriorators": deteriorators,
        "top_improvers": improvers,
    }


async def _trend_meta(
    db,
    cockpit: Dict[str, Any],
    *,
    entity_code: Optional[str],
    period_ym: Optional[str],
    department_id: Optional[str],
    cost_center_id: Optional[str],
    process: Optional[str],
) -> Tuple[str, Optional[str]]:
    stored = await kss.fetch_weekly_trends(
        db,
        entity_code=entity_code,
        period_ym=period_ym,
        department_id=department_id,
        cost_center_id=cost_center_id,
        process=process,
    )
    if len(stored) >= 2:
        return "snapshots", None
    trends = cockpit.get("trends") or []
    if trends and all(t.get("readiness") is not None for t in trends[:2]):
        return "cockpit", None
    return "synthetic", "Series estimated from current readiness when historical snapshots are sparse."


def _readiness_alerts(kpis: Dict[str, Any]) -> List[Dict[str, Any]]:
    from app.services.cfo_command_center_service import _build_alerts

    codes = (
        "audit_readiness_pct",
        "repeat_finding_rate_pct",
        "high_critical_open_cases",
        "remediation_sla_pct",
        "evidence_completeness_pct",
    )
    out = []
    for a in _build_alerts(kpis, []):
        if a.get("code") not in codes:
            continue
        code = a.get("code")
        href = READINESS_ALERT_HREFS.get(code)
        out.append({**a, "href": href} if href else a)
    return out


def _paginate(items: List[Any], limit: int, offset: int) -> Dict[str, Any]:
    total = len(items)
    lim = max(1, min(int(limit or 50), 200))
    off = max(0, int(offset or 0))
    return {
        "items": items[off : off + lim],
        "total": total,
        "limit": lim,
        "offset": off,
        "has_more": off + lim < total,
    }


def _waterfall_steps(portfolio: Dict[str, Any], current: float) -> List[Dict[str, Any]]:
    """Static waterfall from current component gaps vs perfect score (100%)."""
    weights = portfolio.get("weights") or {"control": 40, "recon": 25, "evidence": 20, "issues": 15}
    steps: List[Dict[str, Any]] = [{"label": "Perfect score", "value": 100.0, "kind": "start"}]
    running = 100.0
    for key, label, pct_key in (
        ("control", "Controls (40%)", "control_pct"),
        ("recon", "Recon (25%)", "recon_pct"),
        ("evidence", "Evidence (20%)", "evidence_pct"),
        ("issues", "Issues (15%)", "issue_pct"),
    ):
        w = float(weights.get(key) or 0)
        gap = round(w * (100.0 - float(portfolio.get(pct_key) or 0)) / 100.0, 1)
        if gap > 0.05:
            running = round(running - gap, 1)
            steps.append({"label": f"− {label}", "value": running, "delta": -gap, "kind": "debit"})
    steps.append({"label": "Portfolio readiness", "value": round(float(current or 0), 1), "kind": "end"})
    return steps


def _risk_band_for_readiness(summary: Dict[str, Any], kpis: Dict[str, Any]) -> str:
    from app.services.cfo_executive_narrative_ml import _composite_risk_score, _norm_count, _norm_log_usd, _norm_pct

    cur = float(summary.get("current") or 0)
    features = {
        "readiness_gap": _norm_pct(100.0 - cur, invert=False),
        "exposure_norm": _norm_log_usd(kpis.get("unresolved_high_risk_exposure")),
        "cases_norm": _norm_count(kpis.get("high_critical_open_cases")),
        "repeat_norm": _norm_pct(kpis.get("repeat_finding_rate_pct"), invert=False),
        "evidence_gap": _norm_pct(100.0 - float(kpis.get("evidence_completeness_pct") or 80), invert=False),
        "sla_gap": _norm_pct(100.0 - float(kpis.get("remediation_sla_pct") or 90), invert=False),
        "trend_readiness_down": 0.0,
        "trend_exposure_up": 0.0,
        "alert_density": 0.0,
        "ops_pressure": 0.0,
    }
    return _risk_band_from_score(_composite_risk_score(features))


def _risk_band_from_score(score: float) -> str:
    if score >= 0.75:
        return "critical"
    if score >= 0.55:
        return "elevated"
    if score >= 0.35:
        return "moderate"
    return "stable"


async def _extra_metrics(
    db,
    heatmap: List[Dict[str, Any]],
    cockpit: Dict[str, Any],
    *,
    entity_code: Optional[str],
    period_ym: Optional[str],
    department_id: Optional[str],
    cost_center_id: Optional[str],
    process: Optional[str],
) -> Dict[str, Any]:
    """Portfolio-level driver metrics for drill-down explainability."""
    if heatmap:
        global_control_pass_pct = round(
            sum(float(r.get("control_component") or 0) for r in heatmap) / len(heatmap) * 100,
            1,
        )
    else:
        global_control_pass_pct = 0.0

    rq: Dict[str, Any] = {"status": "overdue"}
    if entity_code:
        rq["entity"] = entity_code
    if period_ym:
        rq["period"] = period_ym
    overdue_recons = await db.reconciliations.count_documents(rq)
    overdue_items: List[Dict[str, Any]] = []
    async for r in db.reconciliations.find(rq, {"_id": 0}).sort("id", 1).limit(25):
        overdue_items.append(
            {
                "id": r.get("id"),
                "entity": r.get("entity"),
                "period": r.get("period"),
                "account": r.get("account") or r.get("name"),
                "status": r.get("status"),
            }
        )

    cq = merge_cases_master_filters(
        {"status": {"$ne": "closed"}, "closed_at": None},
        entity_code=entity_code,
        period_ym=period_ym,
        department_id=department_id,
        cost_center_id=cost_center_id,
    )
    if process:
        cq["process"] = process
    cases_no_evidence: List[Dict[str, Any]] = []
    async for c in db.cases.find(cq, {"_id": 0, "id": 1, "title": 1, "entity": 1, "process": 1, "severity": 1}).limit(25):
        cases_no_evidence.append(
            {
                "id": c.get("id"),
                "title": c.get("title"),
                "entity": c.get("entity"),
                "process": c.get("process"),
                "severity": c.get("severity"),
            }
        )
    cases_no_evidence_count = await db.cases.count_documents(cq)

    ctrl_q: Dict[str, Any] = {"last_run_pass": None}
    if process:
        ctrl_q["process"] = process
    controls_never_run: List[Dict[str, Any]] = []
    async for c in db.controls.find(ctrl_q, {"_id": 0, "code": 1, "name": 1, "process": 1}).sort("code", 1).limit(25):
        controls_never_run.append({"code": c.get("code"), "name": c.get("name"), "process": c.get("process")})
    never_run_count = await db.controls.count_documents(ctrl_q)

    per_control: Dict[str, int] = defaultdict(int)
    ex_q = _scope_exceptions(
        None,
        entity_code=entity_code,
        period_ym=period_ym,
        department_id=department_id,
        cost_center_id=cost_center_id,
        process=process,
    )
    async for ex in db.exceptions.find(ex_q if ex_q else {}, {"_id": 0, "control_code": 1}):
        code = ex.get("control_code")
        if code:
            per_control[code] += 1
    repeat_offenders = [
        {"code": code, "exceptions": cnt}
        for code, cnt in sorted(per_control.items(), key=lambda kv: -kv[1])
        if cnt > 1
    ][:15]

    return {
        "global_control_pass_pct": global_control_pass_pct,
        "overdue_reconciliations_count": overdue_recons,
        "overdue_reconciliations": overdue_items,
        "cases_without_evidence_count": cases_no_evidence_count,
        "cases_without_evidence": cases_no_evidence,
        "controls_never_run_count": never_run_count,
        "controls_never_run": controls_never_run,
        "repeat_offenders": repeat_offenders,
    }


def _narrative_slice(cockpit: Dict[str, Any], alerts: List[Dict[str, Any]]) -> Dict[str, Any]:
    from app.services.cfo_executive_narrative_ml import build_executive_narrative_ml

    fa = cockpit.get("filters_applied") or {}
    narrative = build_executive_narrative_ml(
        cockpit=cockpit,
        alerts=alerts,
        ops_kpis=cockpit.get("ops_kpis"),
        entity_code=fa.get("entity_code"),
        period_ym=fa.get("period_ym"),
        process=fa.get("process"),
    )
    sections = narrative.get("sections") or {}
    return {
        "model": narrative.get("model"),
        "risk_band": sections.get("risk_band"),
        "executive_snapshot": sections.get("summary") or sections.get("executive_snapshot"),
        "drivers": sections.get("drivers") or [],
        "recommended_actions": sections.get("recommended_actions") or sections.get("actions") or [],
        "citations": (narrative.get("citations") or [])[:6],
    }


def _recommended_actions(summary: Dict[str, Any], weakest: Optional[Dict[str, Any]], alerts: List[Dict[str, Any]]) -> List[str]:
    actions: List[str] = []
    cur = summary.get("current")
    if isinstance(cur, (int, float)) and cur < READINESS_WARN_PCT:
        actions.append(f"Prioritize readiness recovery: portfolio at {cur}% (target {READINESS_TARGET_PCT:.0f}%).")
    if weakest:
        actions.append(
            f"Focus on {weakest.get('process')} @ {weakest.get('entity')} "
            f"({weakest.get('readiness')}% · {weakest.get('open_high', 0)} open high/critical)."
        )
    if summary.get("cells_below_60", 0) > 0:
        actions.append(f"Remediate {summary['cells_below_60']} entity×process cells below 60% before period close.")
    for a in alerts[:2]:
        if a.get("message"):
            actions.append(str(a["message"]))
    if summary.get("p0_open", 0) > 0:
        actions.append(f"Clear {summary['p0_open']} P0 action-queue items linked to assurance processes.")
    return actions[:5]


async def build_audit_readiness_detail(
    db,
    cockpit: Dict[str, Any],
    *,
    entity_code: Optional[str] = None,
    period_ym: Optional[str] = None,
    department_id: Optional[str] = None,
    cost_center_id: Optional[str] = None,
    process: Optional[str] = None,
    cells_limit: int = 50,
    cells_offset: int = 0,
    controls_limit: int = 10,
    controls_offset: int = 0,
    exceptions_limit: int = 15,
    exceptions_offset: int = 0,
) -> Dict[str, Any]:
    heatmap = list(cockpit.get("heatmap") or [])
    kpis = dict(cockpit.get("kpis") or {})
    current = float(kpis.get("audit_readiness_pct") or 0)

    scope = kss.scope_key(
        entity_code=entity_code,
        period_ym=period_ym,
        department_id=department_id,
        cost_center_id=cost_center_id,
        process=process,
    )
    prior_doc = await kss.fetch_prior_snapshot(db, scope, skip_latest=1)
    prior_kpis = (prior_doc or {}).get("kpis") or {}
    prior_cells = (prior_doc or {}).get("readiness_cells") or {}
    prior_readiness = prior_kpis.get("audit_readiness_pct")
    delta = kss._compute_delta(current, prior_readiness, "pct")

    portfolio = _portfolio_components(heatmap)
    await kss.record_snapshot(
        db,
        kpis,
        entity_code=entity_code,
        period_ym=period_ym,
        department_id=department_id,
        cost_center_id=cost_center_id,
        process=process,
        readiness_cells=kss.heatmap_to_cell_map(heatmap),
        portfolio_components=kss.portfolio_components_snapshot(portfolio),
        min_interval_hours=4.0,
    )

    from app.services import action_queue_analytics_service as aqa

    aq_summary = await aqa.build_summary(
        db,
        entity_code=entity_code,
        period_ym=period_ym,
        department_id=department_id,
        cost_center_id=cost_center_id,
        process=process,
    )
    trend_source, trend_note = await _trend_meta(
        db,
        cockpit,
        entity_code=entity_code,
        period_ym=period_ym,
        department_id=department_id,
        cost_center_id=cost_center_id,
        process=process,
    )

    cells_below_60 = sum(1 for r in heatmap if float(r.get("readiness") or 0) < 60)
    cells_above_80 = sum(1 for r in heatmap if float(r.get("readiness") or 0) >= 80)
    total_exposure = round(sum(float(r.get("exposure") or 0) for r in heatmap), 2)
    open_high_total = sum(int(r.get("open_high") or 0) for r in heatmap)

    weakest = _weakest_cell(heatmap)
    risk_band = _risk_band_for_readiness({"current": current}, kpis)
    summary = {
        "current": current,
        "prior_value": delta.get("prior_value"),
        "delta_pct": delta.get("delta_pct"),
        "delta_abs": delta.get("delta_abs"),
        "delta_direction": delta.get("delta_direction"),
        "severity": _severity_for_pct(current, good=READINESS_TARGET_PCT, warn=READINESS_WARN_PCT),
        "risk_band": risk_band,
        "target_pct": READINESS_TARGET_PCT,
        "gap_to_target": round(READINESS_TARGET_PCT - current, 1),
        "cell_count": len(heatmap),
        "cells_below_60": cells_below_60,
        "cells_above_80": cells_above_80,
        "total_exposure_usd": total_exposure,
        "open_high_total": open_high_total,
        "p0_open": int(aq_summary.get("p0_open") or 0),
        "p1_open": int(aq_summary.get("p1_open") or 0),
    }

    correlated = []
    for kid in CORRELATED_KPI_IDS:
        val = kpis.get(kid)
        if val is None:
            continue
        correlated.append({"id": kid, "value": val, "unit": "usd" if "exposure" in kid else ("pct" if "pct" in kid else "count")})

    alerts = _readiness_alerts(kpis)
    movers = _movers_from_heatmap(heatmap, prior_cells if prior_cells else None)
    weak_cells = [r for r in heatmap if float(r.get("readiness") or 100) < 60]
    weak_cells.sort(key=lambda r: float(r.get("readiness") or 0))
    failing = list(cockpit.get("top_failing_controls") or [])
    risks = list(cockpit.get("top_risks") or [])

    extra = await _extra_metrics(
        db,
        heatmap,
        cockpit,
        entity_code=entity_code,
        period_ym=period_ym,
        department_id=department_id,
        cost_center_id=cost_center_id,
        process=process,
    )
    narrative_slice = _narrative_slice(cockpit, alerts) if readiness_drill_v2_enabled() else None

    return {
        "summary": summary,
        "heatmap": heatmap,
        "top_failing_controls": failing,
        "top_risks": risks,
        "portfolio_components": portfolio,
        "distribution": _distribution(heatmap),
        "waterfall": _waterfall_steps(portfolio, current),
        "weakest_cell": weakest,
        "weakest_entity": _weakest_by_key(heatmap, "entity"),
        "weakest_process": _weakest_by_key(heatmap, "process"),
        "correlated_kpis": correlated,
        "movers": movers,
        "alerts": alerts,
        "extra_metrics": extra,
        "narrative_slice": narrative_slice,
        "recommended_actions": _recommended_actions(summary, weakest, alerts),
        "lists": {
            "weak_cells": _paginate(weak_cells, cells_limit, cells_offset),
            "failing_controls": _paginate(failing, controls_limit, controls_offset),
            "open_exceptions": _paginate(risks, exceptions_limit, exceptions_offset),
        },
        "committee_pack_path": "/api/reports/audit-committee-pack.xlsx",
        "trend_source": trend_source,
        "trend_note": trend_note,
        "target_pct": READINESS_TARGET_PCT,
        "filters_applied": cockpit.get("filters_applied") or {},
        "as_of": _as_of_now(),
        "feature_flags": {"readiness_drill_v2": readiness_drill_v2_enabled()},
    }


async def audit_readiness_trend(
    db,
    *,
    entity_code: Optional[str] = None,
    period_ym: Optional[str] = None,
    department_id: Optional[str] = None,
    cost_center_id: Optional[str] = None,
    process: Optional[str] = None,
) -> Dict[str, Any]:
    cockpit = await cfo_cockpit(
        db,
        entity_code=entity_code,
        period_ym=period_ym,
        department_id=department_id,
        cost_center_id=cost_center_id,
        process=process,
    )
    trends = cockpit.get("trends") or []
    trend_source, note = await _trend_meta(
        db,
        cockpit,
        entity_code=entity_code,
        period_ym=period_ym,
        department_id=department_id,
        cost_center_id=cost_center_id,
        process=process,
    )

    series = []
    multi = []
    for t in trends:
        period = t.get("week") or ""
        if not period:
            continue
        readiness = round(float(t.get("readiness") or 0), 1)
        series.append({"period": period, "value": readiness})
        multi.append(
            {
                "period": period,
                "readiness": readiness,
                "exposure": round(float(t.get("exposure") or 0), 2),
                "control_fail_count": int(t.get("control_fail_count") or 0),
            }
        )

    prior_series = []
    scope = kss.scope_key(
        entity_code=entity_code,
        period_ym=period_ym,
        department_id=department_id,
        cost_center_id=cost_center_id,
        process=process,
    )
    prior_doc = await kss.fetch_prior_snapshot(db, scope, skip_latest=2)
    if prior_doc:
        pk = (prior_doc.get("kpis") or {}).get("audit_readiness_pct")
        if pk is not None and series:
            prior_series = [{"period": "prior_period", "value": round(float(pk), 1)}]

    component_series = await kss.fetch_component_trends(
        db,
        entity_code=entity_code,
        period_ym=period_ym,
        department_id=department_id,
        cost_center_id=cost_center_id,
        process=process,
    )

    return {
        "kpi_id": "audit_readiness_pct",
        "series": series,
        "multi_series": multi,
        "component_series": component_series,
        "prior_anchor": prior_series,
        "value_kind": "pct",
        "target_pct": READINESS_TARGET_PCT,
        "trend_source": trend_source,
        "note": note,
        "source": "dashboard/cfo",
        "as_of": _as_of_now(),
    }


def export_xlsx_bytes(detail: Dict[str, Any]) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    summary = detail.get("summary") or {}
    for row in (
        ("As of", detail.get("as_of")),
        ("Current %", summary.get("current")),
        ("Prior %", summary.get("prior_value")),
        ("Delta pts", summary.get("delta_pct")),
        ("Risk band", summary.get("risk_band")),
        ("Cells below 60%", summary.get("cells_below_60")),
        ("Cells above 80%", summary.get("cells_above_80")),
    ):
        ws_sum.append(list(row))

    ws_hm = wb.create_sheet("Heatmap")
    ws_hm.append(
        ["Entity", "Process", "Readiness", "Control%", "Recon%", "Evidence%", "Issues%", "Open high", "Exposure"]
    )
    for r in detail.get("heatmap") or []:
        ws_hm.append(
            [
                r.get("entity"),
                r.get("process"),
                r.get("readiness"),
                round(float(r.get("control_component") or 0) * 100, 1),
                round(float(r.get("recon_component") or 0) * 100, 1),
                round(float(r.get("evidence_component") or 0) * 100, 1),
                round(float(r.get("issue_component") or 0) * 100, 1),
                r.get("open_high"),
                r.get("exposure"),
            ]
        )

    extra = detail.get("extra_metrics") or {}
    if extra:
        ws_ex = wb.create_sheet("Drivers")
        ws_ex.append(["Metric", "Value"])
        ws_ex.append(["Global control pass %", extra.get("global_control_pass_pct")])
        ws_ex.append(["Overdue recons", extra.get("overdue_reconciliations_count")])
        ws_ex.append(["Cases w/o evidence", extra.get("cases_without_evidence_count")])
        ws_ex.append(["Controls never run", extra.get("controls_never_run_count")])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_csv_bytes(detail: Dict[str, Any]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    summary = detail.get("summary") or {}
    w.writerow(["Audit readiness export", detail.get("as_of", "")])
    w.writerow(["Current %", summary.get("current")])
    w.writerow(["Prior %", summary.get("prior_value")])
    w.writerow(["Delta pts", summary.get("delta_pct")])
    w.writerow([])
    w.writerow(["Entity", "Process", "Readiness", "Control%", "Recon%", "Evidence%", "Issues%", "Open high", "Exposure"])
    for r in detail.get("heatmap") or []:
        w.writerow(
            [
                r.get("entity"),
                r.get("process"),
                r.get("readiness"),
                round(float(r.get("control_component") or 0) * 100, 1),
                round(float(r.get("recon_component") or 0) * 100, 1),
                round(float(r.get("evidence_component") or 0) * 100, 1),
                round(float(r.get("issue_component") or 0) * 100, 1),
                r.get("open_high"),
                r.get("exposure"),
            ]
        )
    return buf.getvalue().encode("utf-8")
