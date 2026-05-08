"""Generate Audit Committee pack as PDF + workbook as XLSX.

Both rendered in-memory, returned as bytes for HTTP streaming.
"""
from __future__ import annotations
import io
from datetime import datetime, timezone
from typing import Any, Dict, Optional

try:
    # Optional dependency in some dev environments; API should fail gracefully if not installed.
    from reportlab.lib.pagesizes import A4  # type: ignore
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # type: ignore
    from reportlab.lib import colors  # type: ignore
    from reportlab.lib.units import cm  # type: ignore
    from reportlab.platypus import (  # type: ignore
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    )
except ModuleNotFoundError:  # pragma: no cover
    A4 = None  # type: ignore
    getSampleStyleSheet = None  # type: ignore
    ParagraphStyle = None  # type: ignore
    colors = None  # type: ignore
    cm = None  # type: ignore
    SimpleDocTemplate = None  # type: ignore
    Paragraph = None  # type: ignore
    Spacer = None  # type: ignore
    Table = None  # type: ignore
    TableStyle = None  # type: ignore
    PageBreak = None  # type: ignore

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .analytics import cfo_cockpit


def _now_str():
    return datetime.now(timezone.utc).strftime("%d %b %Y %H:%M UTC")


async def build_pdf(
    db,
    *,
    entity_code: Optional[str] = None,
    period_ym: Optional[str] = None,
    department_id: Optional[str] = None,
    cost_center_id: Optional[str] = None,
) -> bytes:
    if A4 is None:
        raise RuntimeError("PDF export requires optional dependency 'reportlab' to be installed")
    data = await cfo_cockpit(
        db,
        entity_code=entity_code,
        period_ym=period_ym,
        department_id=department_id,
        cost_center_id=cost_center_id,
    )
    k = data["kpis"]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm,
                            leftMargin=2 * cm, rightMargin=2 * cm, title="Audit Committee Pack")
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Mono", fontName="Courier", fontSize=8, textColor=colors.grey, spaceAfter=4))
    styles.add(ParagraphStyle(name="HeroTitle", fontName="Helvetica-Bold", fontSize=28, textColor=colors.black, leading=30, spaceAfter=10))
    styles.add(ParagraphStyle(name="SectionTitle", fontName="Helvetica-Bold", fontSize=14, textColor=colors.black, spaceBefore=18, spaceAfter=10))
    styles.add(ParagraphStyle(name="Body", fontName="Helvetica", fontSize=10, textColor=colors.HexColor("#333333"), leading=14, spaceAfter=8))

    story = []
    story.append(Paragraph("ONE TOUCH AUDIT AI", styles["Mono"]))
    story.append(Paragraph("Audit Committee Pack", styles["HeroTitle"]))
    story.append(Paragraph(f"Generated {_now_str()}", styles["Mono"]))
    story.append(Spacer(1, 12))
    fa = data.get("filters_applied") or {}
    ctx = ", ".join(f"{k}={v}" for k, v in sorted(fa.items())) if fa else "Enterprise (no entity / period / dept / CC filter)"
    story.append(Paragraph(f"<b>Reporting context (Phase 13):</b> {ctx}", styles["Body"]))
    story.append(Spacer(1, 12))

    # KPI summary
    story.append(Paragraph("Executive summary", styles["SectionTitle"]))
    kpi_rows = [
        ["Audit readiness %", f"{k['audit_readiness_pct']:.1f}%"],
        ["Unresolved high-risk exposure", f"${k['unresolved_high_risk_exposure']:,.2f}"],
        ["High/critical open cases", f"{k['high_critical_open_cases']}"],
        ["Total open cases", f"{k['open_cases']}"],
        ["Repeat finding rate %", f"{k['repeat_finding_rate_pct']:.1f}%"],
        ["Evidence completeness %", f"{k['evidence_completeness_pct']:.1f}%"],
        ["Remediation SLA %", f"{k['remediation_sla_pct']:.1f}%"],
    ]
    t = Table(kpi_rows, colWidths=[8 * cm, 6 * cm])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.HexColor("#AAAAAA")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("TEXTCOLOR", (1, 0), (1, -1), colors.black),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (1, 0), (1, -1), "Courier-Bold"),
    ]))
    story.append(t)

    # Top failing controls
    story.append(Paragraph("Top failing controls", styles["SectionTitle"]))
    tf_rows = [["Code", "Control", "Process", "Criticality", "Exceptions"]]
    for c in data["top_failing_controls"]:
        tf_rows.append([c["code"], c["name"], c["process"], c["criticality"], str(c["exceptions"])])
    tt = Table(tf_rows, colWidths=[2.4 * cm, 6.5 * cm, 3.5 * cm, 2.5 * cm, 2 * cm])
    tt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#222222")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#DDDDDD")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("ALIGN", (4, 1), (4, -1), "RIGHT"),
    ]))
    story.append(tt)

    # Top risks
    story.append(PageBreak())
    story.append(Paragraph("Top unresolved risks", styles["SectionTitle"]))
    risk_rows = [["Issue", "Control", "Entity", "Severity", "Exposure"]]
    for r in data["top_risks"]:
        risk_rows.append([
            (r["title"][:60] + "…") if len(r["title"]) > 60 else r["title"],
            r["control_code"], r["entity"], r["severity"].upper(),
            f"${r['financial_exposure']:,.0f}"
        ])
    rt = Table(risk_rows, colWidths=[8 * cm, 2.4 * cm, 2 * cm, 2 * cm, 2.6 * cm])
    rt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#222222")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#DDDDDD")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("ALIGN", (4, 1), (4, -1), "RIGHT"),
    ]))
    story.append(rt)

    # Process × entity heatmap table
    story.append(PageBreak())
    story.append(Paragraph("Process × entity readiness heatmap", styles["SectionTitle"]))
    entities = sorted({r["entity"] for r in data["heatmap"]})
    processes = sorted({r["process"] for r in data["heatmap"]})
    rows = [["Process / Entity"] + entities]
    lookup = {(r["entity"], r["process"]): r for r in data["heatmap"]}
    for p in processes:
        row = [p]
        for e in entities:
            cell = lookup.get((e, p))
            row.append(f"{cell['readiness']:.0f}" if cell else "—")
        rows.append(row)
    ht = Table(rows, colWidths=[4 * cm] + [2.5 * cm] * len(entities))
    hts = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#222222")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.HexColor("#DDDDDD")),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
    ]
    # color cells
    for i in range(1, len(rows)):
        for j in range(1, len(rows[0])):
            v_str = rows[i][j]
            try:
                v = float(v_str)
                if v >= 80: col = colors.HexColor("#D8F3DC")
                elif v >= 65: col = colors.HexColor("#FFF3BF")
                elif v >= 50: col = colors.HexColor("#FFD6A5")
                else: col = colors.HexColor("#FFADAD")
                hts.append(("BACKGROUND", (j, i), (j, i), col))
            except ValueError:
                pass
    ht.setStyle(TableStyle(hts))
    story.append(ht)

    story.append(Spacer(1, 24))
    story.append(Paragraph("Prepared by One Touch Audit AI · Continuous Assurance Platform. This pack is an automated extract; material conclusions require human review.", styles["Mono"]))

    doc.build(story)
    return buf.getvalue()


async def build_xlsx(
    db,
    *,
    entity_code: Optional[str] = None,
    period_ym: Optional[str] = None,
    department_id: Optional[str] = None,
    cost_center_id: Optional[str] = None,
) -> bytes:
    data = await cfo_cockpit(
        db,
        entity_code=entity_code,
        period_ym=period_ym,
        department_id=department_id,
        cost_center_id=cost_center_id,
    )
    k = data["kpis"]
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="111111")
    header_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    border = Border(*[Side(style="thin", color="CCCCCC")] * 4)

    # 1. Executive Summary
    ws = wb.active
    ws.title = "Executive Summary"
    fa = data.get("filters_applied") or {}
    ctx = ", ".join(f"{k}={v}" for k, v in sorted(fa.items())) if fa else "Enterprise (no entity / period / dept / CC filter)"
    ws.append(["Reporting context (Phase 13)", ctx])
    ws.append([])
    ws.append(["KPI", "Value"])
    header_row_idx = 3
    for cell in ws[header_row_idx]:
        cell.fill = header_fill
        cell.font = header_font
    for row in [
        ["Audit readiness %", round(k["audit_readiness_pct"], 1)],
        ["Unresolved high-risk exposure (USD)", round(k["unresolved_high_risk_exposure"], 2)],
        ["High/critical open cases", k["high_critical_open_cases"]],
        ["Total open cases", k["open_cases"]],
        ["Repeat finding rate %", round(k["repeat_finding_rate_pct"], 1)],
        ["Evidence completeness %", round(k["evidence_completeness_pct"], 1)],
        ["Remediation SLA %", round(k["remediation_sla_pct"], 1)],
    ]:
        ws.append(row)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18

    # 2. Top Risks
    ws2 = wb.create_sheet("Top Risks")
    headers = ["Title", "Control", "Entity", "Severity", "Exposure (USD)", "Anomaly", "Materiality"]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.fill = header_fill; cell.font = header_font
    for r in data["top_risks"]:
        ws2.append([
            r["title"], r["control_code"], r["entity"], r["severity"].upper(),
            round(r["financial_exposure"], 2), r["anomaly_score"], r["materiality_score"]
        ])
    for i, w in enumerate([60, 12, 10, 10, 16, 10, 12], start=1):
        ws2.column_dimensions[chr(ord("A") + i - 1)].width = w

    # 3. Top Failing Controls
    ws3 = wb.create_sheet("Top Failing Controls")
    ws3.append(["Code", "Control", "Process", "Criticality", "Exceptions"])
    for cell in ws3[1]:
        cell.fill = header_fill; cell.font = header_font
    for c in data["top_failing_controls"]:
        ws3.append([c["code"], c["name"], c["process"], c["criticality"], c["exceptions"]])

    # 4. Heatmap
    ws4 = wb.create_sheet("Readiness Heatmap")
    entities = sorted({r["entity"] for r in data["heatmap"]})
    processes = sorted({r["process"] for r in data["heatmap"]})
    ws4.append(["Process / Entity"] + entities)
    for cell in ws4[1]:
        cell.fill = header_fill; cell.font = header_font
    lookup = {(r["entity"], r["process"]): r for r in data["heatmap"]}
    for p in processes:
        row = [p]
        for e in entities:
            cell = lookup.get((e, p))
            row.append(round(cell["readiness"], 1) if cell else None)
        ws4.append(row)

    # Autosize-ish for headers
    for sh in wb.worksheets:
        for col_cells in sh.columns:
            max_len = 0
            for c in col_cells:
                v = c.value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            letter = col_cells[0].column_letter
            sh.column_dimensions[letter].width = max(sh.column_dimensions[letter].width or 10, min(60, max_len + 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
